import csv
import openpyxl
import os
from pathlib import Path
import re

regex = re.compile('^.*.xlsx$')

data = []
path = Path.cwd()/'excelSpreadsheets'


for filename in os.listdir(path):
    if regex.search(filename) :

        file = openpyxl.load_workbook(path/filename)
        for sheetname in file.sheetnames:
            sheet = file[sheetname]
            fp = open(f'csvs/{filename[:-5]}_{sheetname}.csv','w',newline='')
            Writer = csv.writer(fp)
            for i in range(1,sheet.max_row+1):
                rowdata = []
                for j in range(1,sheet.max_column+1):
                    rowdata.append(sheet.cell(i,j).value)
                Writer.writerow(rowdata)
            fp.close()

