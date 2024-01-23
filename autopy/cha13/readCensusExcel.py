
# rreadCensusExcel - Tabulates population and number of census tracts for
# each country.

import openpyxl,pprint
print('Openning workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']

countyData = {}

# TODO: Fill in countyData with county's populatioin and tracts.
print('Reading now...')
for row in range(2,sheet.max_row+1):
    state = sheet['B'+str(row)].value
    county = sheet.cell(row,3).value
    pop = sheet.cell(row,4).value
    # TODO: Open a new file and write county data to it
    # if state in countyData.keys():
    #     statepop = countyData[state]
    #     if county in statepop.keys():
    #         countypop = statepop[county]
    #         countypop['tracts'] += 1
    #         countypop['pop'] += pop
    #     else:
    #         statepop[county] = {'tracts':1,'pop':pop}
    # else:
    #     countyData[state] = {}
    countyData.setdefault(state,{})
    countyData[state].setdefault(county,{'tracts':0,'pop':0})
    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += pop 

print('writing results...')
fp = open('census2010.py','w')
fp.write('allDate = '+pprint.pformat(countyData))
fp.close()
print('Done.')

