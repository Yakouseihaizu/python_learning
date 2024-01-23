# import openpyxl
# wb = openpyxl.load_workbook('example.xlsx')
# # print(type(wb))
# print(wb.sheetnames)
# sheet = wb['Sheet3']
# print(wb)
# print(sheet)

# print(type(sheet))
# print(sheet.title)
# anothorSheet = wb.active
# print(anothorSheet)
# print(anothorSheet.title)

# import openpyxl
# wb = openpyxl.load_workbook(r'excel\example.xlsx') #不行，只能在当前文件夹
# sheet = wb['Sheet1']
# for row_of_objects in sheet['A1':'C3']:
#     for obj in row_of_objects:
#         print(obj.coordinate,obj.value)
#     print('-----END-OF-ROW-----')

# import openpyxl
# wb = openpyxl.Workbook()
# sheet = wb.active
# print(wb.sheetnames)
# sheet.title = 'spam'
# print(wb.sheetnames)
# wb.save('new.xlsx')

# import openpyxl
# import logging
# logging.basicConfig(format='%(levelname)s - %(message)s',level=logging.DEBUG)

# wb = openpyxl.Workbook()
# # logging.debug(str(wb.sheetnames))
# # wb.create_sheet('fisrt sheet',0)
# # logging.debug(str(wb.sheetnames))
# # wb.create_sheet('mid sheet',1)
# # logging.debug(str(wb.sheetnames))
# # del wb['Sheet']
# # logging.debug(str(wb.sheetnames))
# # wb.save('add_delet_sheet.xlsx')
# # sheet = wb.active
# # openpyxl.Workbook
# # sheet['A1'] = 'Hello, world!'
# # print(sheet['A1'])
# # print(sheet['A1'].value)

# from openpyxl.styles import Font

# sheet = wb.active
# # sheet['A1'].font = Font(size=24,italic=True)
# # sheet['A1'] = 'Hello,world!'
# # sheet['B1'] = 'Anyway'
# sheet['A1'] = 'Blod Times New Roman'
# sheet['A1'].font = Font('Times New Roman',bold=True)

# wb.save('font.xlsx')

# import openpyxl
# import logging
# logging.basicConfig(format='%(message)s - %(acsdate)s')

# wb = openpyxl.Workbook()
# sheet = wb.active
# # sheet['A1'] = 'Wide column'
# # sheet['B2'] = 'Tall row'
# # sheet.row_dimensions[2].height = 50
# # sheet.column_dimensions['A'].width = 40
# # wb.save('dimension.xlsx')
# sheet['A1'] = 'Tewlve cells merged together.'
# sheet.merge_cells('A1:D3')

# sheet.merge_cells('C5:D5')
# sheet['C5'] = 'Two merged cells.'
# sheet.unmerge_cells('C5:D5')
# # print(sheet['C5'].value)
# wb.save('merged.xlsx')

# import openpyxl

# wb = openpyxl.load_workbook('produceSales.xlsx')
# sheet = wb.active
# sheet.freeze_panes = 'B2'
# wb.save('freezeExmple2.xlsx')

# import openpyxl
# wb = openpyxl.Workbook()
# sheet = wb.active
# for i in range(1,11):
#     sheet['A'+str(i)] = i

# refObj = openpyxl.chart.Reference(sheet,1,10,1,1)
# seriesObj = openpyxl.chart.Series(refObj,title='First series')
# chartObj = openpyxl.chart.BarChart()
# chartObj.title = 'My Chart'
# chartObj.append(seriesObj)

# sheet.add_chart(chartObj,'C5')
# wb.save('sampleChart.xlsx')

import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
ws = wb.active

bars = openpyxl.chart.BarChart()
bars.title = 'sales for fruits'
bars.x_axis.title = 'kinds of fruits'
bars.y_axis.title = 'sales'

data = openpyxl.chart.Reference(ws,min_col=3,max_col=3,min_row=1,max_row=7)
series = openpyxl.chart.Reference(ws,min_col=2,max_col=2,min_row=1,max_row=7)

bars.add_data(data,titles_from_data=False)
bars.set_categories(series)

ws.add_chart(bars,'A10')

wb.save('chart.xlsx')
