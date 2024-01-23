import openpyxl
import sys
from openpyxl.utils import get_column_letter
import datetime
from openpyxl.styles import NamedStyle

import logging
logging.basicConfig(format='%(message)s',level=logging.DEBUG)
logging.disable(logging.DEBUG)

begin = 2# int(sys.argv[1])
lenth = 3#int(sys.argv[2])
filename = 'example.xlsx' #sys.argv[3]

wb = openpyxl.load_workbook(filename)
sheet = wb.active
logging.debug('open '+filename)

max_row = sheet.max_row
# date_time = NamedStyle(name='datetime',number_format='d/m/yy hh/MM')
# print(max_row)
logging.debug(f'got the part to move, including {len(list(sheet.rows)[begin-1])}')
for i in range(max_row-begin+1):
    for j in  range(1,sheet.max_column+1):
        letter = get_column_letter(j)
        content = sheet[letter+str(max_row-i)].value
        sheet[letter+str(max_row+lenth-i)] = content
        logging.debug(f'{letter+str(max_row+lenth-i)} value is {sheet[letter+str(max_row+lenth-i)].value}')
        logging.debug(f'con is {content}')
for i in range(begin,begin+lenth):
    for j in range(1,sheet.max_column+1):
        letter = get_column_letter(j)
        sheet[letter+str(i)] = None

wb.save('new'+filename)