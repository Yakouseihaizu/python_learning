import openpyxl
import sys
from openpyxl.utils import get_column_letter

filename = sys.argv[1]
wb = openpyxl.load_workbook(filename)
sheet = wb.active

max_row = sheet.max_row
max_col = sheet.max_column

data = []

for i in range(1,max_row+1):
    data_in_row = []
    for j in range(1,max_col+1):
        data_in_row.append(sheet.cell(i,j).value)
    data.append(data_in_row)

for i in range(1,max_row+1):
    for j in range(1,max_col+1): 
        letter = get_column_letter(j)
        sheet[letter+str(i)] = None

for i in range(1,max_col+1):
    for j in range(1,max_row+1):
        letter = get_column_letter(j)
        sheet[letter+str(i)] = data[j-1][i-1]

wb.save('new'+filename)

