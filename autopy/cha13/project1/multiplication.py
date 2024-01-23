import sys
import openpyxl
from openpyxl.utils import column_index_from_string,get_column_letter
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active
max = int(sys.argv[1])
for i in range(2,max+2):

    sheet['A'+str(i)].font = Font(bold=True)
    sheet['A'+str(i)] = i-1
    col = get_column_letter(i)

    sheet[col+'1'].font = Font(bold=True)
    sheet[col+'1'] = i-1
    

for i in range(2,max+2):
    for j in range(2,max+2):
        col = get_column_letter(j)
        sheet[col+str(i)] = sheet['A'+str(i)].value*sheet[col+'1'].value

wb.save('multiplcation.xlsx')