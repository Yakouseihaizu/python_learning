import openpyxl
import openpyxl.utils
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']
newdict = {'Lemon':1.27,'Celery':1.9,'Garlic':3.07}

for row in range(2,sheet.max_row+1):
    # if (obj := sheet.cell(row,1)) == 'lemon':
    #     sheet.cell(row,2).value = 1.27
    # elif obj == 'Garlic':
    #     sheet.cell(row,2).value = 3.07
    # elif obj == 'Celery':
    #     sheet.cell(row,2).value = 1.9
    obj = sheet.cell(row,1)
    if obj.value in newdict.keys():
        sheet['B'+str(obj.row)] = newdict[obj.value]
        newdict.pop(obj.value)

for pro,cost in newdict.items():
    row = sheet.max_row+1
    newpro = sheet.cell(row,1)
    newcost = sheet.cell(row,2)
    sheet[newpro.coordinate] = pro
    sheet[newcost.coordinate] = cost
    # ltr = openpyxl.utils.get_column_letter(row)
    # sheet[ltr+str(1)] = pro
    # sheet[ltr+str(2)] = cost


wb.save('updatedSales.xlsx')