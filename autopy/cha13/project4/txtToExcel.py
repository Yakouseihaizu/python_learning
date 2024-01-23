import openpyxl
from pathlib import Path
import os
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active

path = Path.cwd()
texts_num = list(path.glob('*.txt'))
for i in range(len(texts_num)):
    fp = open(texts_num[i],encoding='UTF-8')
    letter = get_column_letter(i+1)
    j=1
    for row in fp.readlines():
        sheet[letter+str(j)] = row
        j+=1
    fp.close()

wb.save('texts2Excel.xlsx')
wb.close()


